import os
from openai import OpenAI
from anthropic import Anthropic
from groq import Groq
#from google.generativeai import GenerativeModel
import google.generativeai as genai
from dotenv import load_dotenv
import openpyxl
import time

load_dotenv()

class TranslationService:
    def __init__(self):
        # Initialize AI providers with API keys from environment variables
        self.openai_client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
        self.anthropic_client = Anthropic(api_key=os.getenv('ANTHROPIC_API_KEY'))
        self.groq_client = Groq(api_key=os.getenv('GROQ_API_KEY'))
        
        # Initialize Google AI
        genai.configure(api_key=os.getenv('GOOGLE_API_KEY'))
        
        # Initialize progress callback
        self.progress_callback = None

    def set_progress_callback(self, callback):
        self.progress_callback = callback

    def update_progress(self, current, total, message=""):
        if self.progress_callback:
            progress = int((current / total) * 100) if total > 0 else 0
            self.progress_callback({
                "progress": progress,
                "message": message
            })

    def translate(self, text, source_lang, target_lang, provider="groq", model=None):
        prompt = f"Translate the following text from {source_lang} to {target_lang}. Preserve any special characters, numbers, and formatting. Only return the translated text, without any explanations:\n\n{text}"
        
        try:
            if provider == "openai":
                response = self.openai_client.chat.completions.create(
                    model=model or "gpt-3.5-turbo",
                    messages=[{"role": "user", "content": prompt}]
                )
                return response.choices[0].message.content.strip()
            
            elif provider == "anthropic":
                response = self.anthropic_client.messages.create(
                    model=model or "claude-3-haiku-20240307",
                    max_tokens=1000,
                    messages=[{"role": "user", "content": prompt}]
                )
                return response.content[0].text.strip()
            
            elif provider == "groq":
                response = self.groq_client.chat.completions.create(
                    model=model or "mixtral-8x7b-32768",
                    messages=[
                        {
                            "role": "system",
                            "content": f"You are a professional translator. Translate from {source_lang} to {target_lang}."
                        },
                        {
                            "role": "user",
                            "content": text
                        }
                    ],
                    temperature=1,
                    max_tokens=1024,
                    top_p=1,
                    stream=False
                )
                return response.choices[0].message.content.strip()
            
            elif provider == "google":
                model = GenerativeModel(model or "gemini-1.5-pro")
                response = model.generate_content(prompt)
                return response.text.strip()
            
            else:
                raise ValueError(f"Unsupported provider: {provider}")
                
        except Exception as e:
            raise Exception(f"Translation error: {str(e)}")

    def translate_excel(self, file_path, target_language, provider="groq", model=None):
        """
        Translate Excel file content while preserving formatting
        """
        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path)
            
            # Calculate total cells that need translation
            total_translatable_cells = 0
            cells_to_translate = []
            
            # First pass: count translatable cells and collect them
            self.update_progress(0, 100, "Analyzing file contents...")
            for sheet in wb.worksheets:
                for row in sheet.rows:
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.strip():
                            total_translatable_cells += 1
                            cells_to_translate.append({
                                'sheet': sheet,
                                'coordinate': cell.coordinate,
                                'text': cell.value,
                                'style': {
                                    'font': cell.font.copy(),
                                    'fill': cell.fill.copy(),
                                    'border': cell.border.copy(),
                                    'alignment': cell.alignment.copy(),
                                    'number_format': cell.number_format
                                }
                            })
            
            if total_translatable_cells == 0:
                self.update_progress(100, 100, "No text to translate!")
                return None
                
            # Initialize progress tracking
            cells_processed = 0
            
            # Process each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_cells = [c for c in cells_to_translate if c['sheet'] == sheet]
                
                if not sheet_cells:
                    continue
                    
                # Get merged cells before unmerging
                merged_ranges = list(sheet.merged_cells.ranges)
                
                # Unmerge cells temporarily
                for merged_range in merged_ranges:
                    sheet.unmerge_cells(str(merged_range))
                
                # Translate cells in this sheet
                for idx, cell_info in enumerate(sheet_cells, 1):
                    try:
                        # Update progress with detailed message
                        progress = int((cells_processed / total_translatable_cells) * 100)
                        self.update_progress(
                            progress,
                            100,
                            f"Translating cell {cells_processed + 1} of {total_translatable_cells} "
                            f"({progress}%) in sheet '{sheet_name}'"
                        )
                        
                        # Translate the cell content
                        translated_text = self.translate(
                            text=cell_info['text'],
                            source_lang="auto",
                            target_lang=target_language,
                            provider=provider,
                            model=model
                        )
                        
                        # Update cell with translation
                        cell = sheet[cell_info['coordinate']]
                        cell.value = translated_text
                        
                        # Restore original styling
                        style = cell_info['style']
                        cell.font = style['font']
                        cell.fill = style['fill']
                        cell.border = style['border']
                        cell.alignment = style['alignment']
                        cell.number_format = style['number_format']
                        
                        # Increment progress
                        cells_processed += 1
                        
                        # Add a small delay to avoid rate limits
                        time.sleep(0.1)
                        
                    except Exception as e:
                        raise Exception(f"Error translating cell {cell_info['coordinate']} in sheet '{sheet_name}': {str(e)}")
                
                # Reapply merged cells
                for merged_range in merged_ranges:
                    sheet.merge_cells(str(merged_range))
            
            # Generate output filename with timestamp
            timestamp = int(time.time())
            output_filename = f"translated_{timestamp}_{os.path.basename(file_path)}"
            output_path = os.path.join("uploads", output_filename)
            
            # Save with progress update
            self.update_progress(95, 100, "Saving translated file...")
            wb.save(output_path)
            
            self.update_progress(100, 100, "Translation completed successfully!")
            return output_filename
            
        except Exception as e:
            self.update_progress(
                cells_processed * 100 // total_translatable_cells if total_translatable_cells > 0 else 0,
                100,
                f"Error: {str(e)}"
            )
            raise Exception(f"Translation failed: {str(e)}")
